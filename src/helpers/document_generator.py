#!/usr/bin/env python3
"""
Document Generator for Hexa RFQ Manager
Generates Excel specification sheets and PDF quote documents
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT


class QuoteDocumentGenerator:
    """Generates quote documents (Excel spec sheet and PDF quote)"""
    
    def __init__(self, specifications: dict, offer_number: str = "41260018"):
        self.specs = specifications
        self.offer_number = offer_number
        self.company_info = {
            'name': 'engionic Femto Gratings GmbH',
            'address': 'Am Stollen 19b',
            'city': '38640 Goslar',
            'country': 'Germany',
            'phone': '+49 30 628873 30',
            'email': 'femto-gratings@engionic.de',
            'website': 'www.engionic-femto-gratings.de',
            'managing_director': 'Tobias Schenk',
            'court': 'Local court Braunschweig',
            'hrb': 'HRB 204454',
            'eori': 'DE764032941655549',
            'tax_no': '2321/208/31038',
            'vat': 'DE815483481',
            'bank': 'Commerzbank AG',
            'iban': 'DE68 1004 0000 0573 6434 00',
            'swift': 'COBADEFFXXX'
        }
        self.pricing = {
            'unit_price': 220.64,
            'quantity': 10,
            'customs': 60.00
        }
        
    def generate_excel_spec_sheet(self, output_path: str) -> str:
        """Generate Excel specification sheet matching the template format"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets in order
        ws_fiber = wb.create_sheet("Fiber Specification")
        ws_sensor = wb.create_sheet("Sensor Specification")
        ws_def = wb.create_sheet("Definitions")
        ws_draw = wb.create_sheet("Drawings")
        
        # Style definitions
        header_font = Font(bold=True, size=11)
        label_font = Font(size=10)
        value_font = Font(size=10, bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        
        # ========== Sensor Specification Sheet ==========
        self._build_sensor_spec_sheet(ws_sensor, header_font, label_font, value_font, border, header_fill)
        
        # ========== Fiber Specification Sheet ==========
        self._build_fiber_spec_sheet(ws_fiber, header_font, label_font, value_font, border, header_fill)
        
        # ========== Definitions Sheet ==========
        self._build_definitions_sheet(ws_def, header_font, label_font, value_font, border)
        
        # ========== Drawings Sheet ==========
        self._build_drawings_sheet(ws_draw, header_font, label_font)
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def _build_sensor_spec_sheet(self, ws, header_font, label_font, value_font, border, header_fill):
        """Build the Sensor Specification sheet"""
        # Header row
        ws['A2'] = 'engionic Fiber Optics GmbH - Ernst-Lau-Straße 8 - 07745 Jena - Germany'
        ws['A2'].font = Font(size=9, color="666666")
        ws['H2'] = 'V1.14'
        ws['H2'].font = Font(size=9, color="666666")
        
        # Production info
        ws['A4'] = 'Production Sheet Number:'
        ws['A4'].font = label_font
        ws['E4'] = 'Position:'
        ws['E4'].font = label_font
        ws['F4'] = 1
        ws['G4'] = 'of:'
        ws['H4'] = 1
        
        ws['A5'] = 'Offer Number:'
        ws['A5'].font = label_font
        ws['C5'] = int(self.offer_number)
        ws['C5'].font = value_font
        ws['E5'] = 'Customer:'
        ws['E5'].font = label_font
        ws['F5'] = self.specs.get('Customer', 'NRL').split(' - ')[0] if ' - ' in self.specs.get('Customer', '') else self.specs.get('Customer', 'NRL')
        ws['F5'].font = value_font
        
        ws['E6'] = 'Quantity:'
        ws['E6'].font = label_font
        ws['F6'] = int(self.specs.get('Quantity', '10').replace(' pcs', '').replace(' pieces', ''))
        ws['F6'].font = value_font
        
        ws['A8'] = 'Contact:'
        ws['A8'].font = label_font
        
        ws['A10'] = 'Remarks FBGs:'
        ws['A10'].font = label_font
        ws['B10'] = 'Kein Faserwechsel ohne Absprache; Toleranz erstes FBG +/-2mm'
        
        ws['A12'] = 'Remarks Sensor:'
        ws['A12'].font = label_font
        
        # Sensor Configuration section
        ws['A14'] = 'Sensor Configuration'
        ws['A14'].font = header_font
        ws['A14'].fill = header_fill
        
        ws['A16'] = 'Sensor Configuration'
        ws['A16'].font = label_font
        ws['B16'] = 'Connector Type'
        ws['B16'].font = label_font
        ws['C16'] = 'Fiber Length (mm)'
        ws['C16'].font = label_font
        
        ws['A18'] = 'Fiber Only'
        ws['A18'].font = value_font
        ws['B18'] = self.specs.get('Connector Type', 'FC/APC both ends')
        ws['B18'].font = value_font
        ws['C18'] = self.specs.get('Total Fiber Length', '10050').replace(' mm', '').replace(',', '')
        ws['C18'].font = value_font
        
        # FBG Specification section
        ws['A25'] = 'FBG Specification'
        ws['A25'].font = header_font
        ws['A25'].fill = header_fill
        
        headers = ['', 'Wavelength (nm)', 'Reflectivity (%)', 'FWHM (nm)', 'min SLSR (dB)', 'FBG Length (mm)', 'FemtoPlus']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=26, column=col, value=header)
            cell.font = label_font
            cell.border = border
        
        # Nominal values
        ws['A27'] = 'Nominal Value'
        ws['B27'] = '-'  # Individual wavelengths in table below
        ws['C27'] = float(self.specs.get('Reflectivity', '10%').replace('%', '').strip()) / 100
        ws['D27'] = float(self.specs.get('FWHM', '0.09').replace(' nm', '').strip())
        ws['E27'] = int(self.specs.get('SLSR Minimum', '8').replace(' dB', '').strip())
        ws['F27'] = int(self.specs.get('FBG Length', '12').replace(' mm', '').strip())
        ws['G27'] = self.specs.get('FemtoPlus', 'Yes')
        
        for col in range(1, 8):
            ws.cell(row=27, column=col).font = value_font
            ws.cell(row=27, column=col).border = border
        
        # Tolerances
        ws['A28'] = 'Plus Tolerance'
        ws['B28'] = float(self.specs.get('Wavelength Tolerance', '±0.1').replace('±', '').replace(' nm', '').strip())
        ws['C28'] = float(self.specs.get('Reflectivity Tolerance', '±4%').replace('±', '').replace('%', '').strip()) / 100
        ws['D28'] = float(self.specs.get('FWHM Tolerance', '±0.02').replace('±', '').replace(' nm', '').strip())
        ws['E28'] = '-'
        ws['F28'] = int(self.specs.get('FBG Length Tolerance', '±2').replace('±', '').replace(' mm', '').strip())
        
        ws['A29'] = 'Minus Tolerance'
        ws['B29'] = float(self.specs.get('Wavelength Tolerance', '±0.1').replace('±', '').replace(' nm', '').strip())
        ws['C29'] = float(self.specs.get('Reflectivity Tolerance', '±4%').replace('±', '').replace('%', '').strip()) / 100
        ws['D29'] = float(self.specs.get('FWHM Tolerance', '±0.02').replace('±', '').replace(' nm', '').strip())
        ws['E29'] = '-'
        ws['F29'] = int(self.specs.get('FBG Length Tolerance', '±2').replace('±', '').replace(' mm', '').strip())
        
        for row in range(28, 30):
            for col in range(1, 7):
                ws.cell(row=row, column=col).font = label_font
                ws.cell(row=row, column=col).border = border
        
        # Additional Specifications
        ws['A32'] = 'Additional Specifications'
        ws['A32'].font = header_font
        ws['A32'].fill = header_fill
        
        ws['B33'] = 'Label'
        ws['C33'] = 'Total Fiber Length (mm)'
        ws['E33'] = 'Tolerance FBG Spacing (%)'
        
        ws['B34'] = self.specs.get('Label', 'on spool')
        ws['C34'] = int(self.specs.get('Total Fiber Length', '10050').replace(' mm', '').replace(',', ''))
        ws['E34'] = 2
        
        # Wavelength Table
        ws['A36'] = 'Wavelength Table'
        ws['A36'].font = header_font
        ws['A36'].fill = header_fill
        
        wl_headers = ['FBG Name', 'Wavelength (nm)', 'FBG Position (mm)', 'FBG Lead Out (mm)', 'FBG Spacing (mm)']
        for col, header in enumerate(wl_headers, 1):
            cell = ws.cell(row=37, column=col, value=header)
            cell.font = label_font
            cell.border = border
        
        # FBG entries
        fbg_wavelengths = [
            self.specs.get('FBG 1 Wavelength', '1550.39 nm').replace(' nm', ''),
            self.specs.get('FBG 2 Wavelength', '1555.39 nm').replace(' nm', '')
        ]
        first_fbg_pos = int(self.specs.get('First FBG Position', '5000').replace(' mm', ''))
        fbg_spacing = int(self.specs.get('FBG Spacing', '50').replace(' mm', ''))
        total_length = int(self.specs.get('Total Fiber Length', '10050').replace(' mm', '').replace(',', ''))
        
        num_fbgs = int(self.specs.get('Number of FBGs', '2'))
        
        for i in range(num_fbgs):
            row = 38 + i
            ws.cell(row=row, column=1, value=f'FBG{i+1}')
            ws.cell(row=row, column=2, value=float(fbg_wavelengths[i]) if i < len(fbg_wavelengths) else '-')
            position = first_fbg_pos + (i * fbg_spacing)
            ws.cell(row=row, column=3, value=position)
            ws.cell(row=row, column=4, value=total_length - position)
            ws.cell(row=row, column=5, value='-' if i == 0 else fbg_spacing)
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).font = value_font
                ws.cell(row=row, column=col).border = border
        
        # Auto Fill configuration (right side)
        ws['G38'] = 'Auto Fill Wavelength Table'
        ws['G39'] = 'Number of FBGs'
        ws['H40'] = num_fbgs
        ws['G42'] = 'Position First FBG (mm)'
        ws['H43'] = first_fbg_pos
        ws['G45'] = 'FBG Wavelength (nm)'
        ws['G46'] = 'Equal Wavelength Spacing'
        ws['G47'] = 'Wavelength Spacing (nm)'
        ws['H48'] = 0
        ws['G49'] = 'Wavelength first FBG'
        ws['H50'] = float(fbg_wavelengths[0])
        ws['G52'] = 'FBG Spacing (mm)'
        ws['G53'] = 'Equal FBG Spacing'
        ws['G54'] = 'Spacing (mm)'
        ws['H55'] = fbg_spacing
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 15
    
    def _build_fiber_spec_sheet(self, ws, header_font, label_font, value_font, border, header_fill):
        """Build the Fiber Specification sheet"""
        # Header
        ws['A2'] = 'engionic Femto Gratings GmbH - Am Stollen 19B - 38640 Goslar'
        ws['A2'].font = Font(size=9, color="666666")
        ws['H2'] = 'V1.14'
        
        # Production info (linked to Sensor Specification)
        ws['A4'] = 'Production Sheet Number:'
        ws['E4'] = 'Position:'
        ws['F4'] = "='Sensor Specification'!F4"
        ws['G4'] = 'of:'
        ws['H4'] = "='Sensor Specification'!H4"
        
        ws['A5'] = 'Offer Number:'
        ws['C5'] = "='Sensor Specification'!C5"
        ws['E5'] = 'Customer:'
        ws['F5'] = "='Sensor Specification'!F5"
        
        ws['A6'] = 'Customer Product ID:'
        ws['E6'] = 'Quantity:'
        ws['F6'] = "='Sensor Specification'!F6"
        
        ws['A8'] = 'Contact:'
        ws['D8'] = 'Approved By:'
        ws['G8'] = 'Date:'
        
        ws['A10'] = 'Remarks:'
        
        # Fiber Specification section
        ws['A12'] = 'Fiber Specification'
        ws['A12'].font = header_font
        ws['A12'].fill = header_fill
        
        fiber_headers = ['Fiber Type', 'Fiber Core Type', 'ITU-T', 'Fiber Cladding Diameter (µm)', 
                        'Fiber Coating Material', 'Fiber Manufacturer', '', 'Fiber Provided By']
        for col, header in enumerate(fiber_headers, 1):
            cell = ws.cell(row=13, column=col, value=header)
            cell.font = label_font
        
        # Fiber data
        ws['A14'] = self.specs.get('Fiber Type', 'SM1330-E9/125PI')
        ws['B14'] = '=VLOOKUP($A$14,Definitions!C2:I16,2,0)'
        ws['C14'] = '=VLOOKUP($A$14,Definitions!C2:I16,7,0)'
        ws['D14'] = '=VLOOKUP($A$14,Definitions!C2:I16,4,0)'
        ws['E14'] = '=VLOOKUP($A$14,Definitions!C2:I16,5,0)'
        ws['F14'] = '=VLOOKUP($A$14,Definitions!C2:I16,6,0)'
        ws['H14'] = 'eFG'
        
        # FBG Specification section
        ws['A16'] = 'FBG Specification'
        ws['A16'].font = header_font
        ws['A16'].fill = header_fill
        
        fbg_headers = ['', 'Wavelength (nm)', 'Reflectivity (%)', 'FWHM (nm)', 
                      'min SLSR (dB)', 'FBG Length (mm)', 'FemtoPlus', 'Apodized']
        for col, header in enumerate(fbg_headers, 1):
            cell = ws.cell(row=17, column=col, value=header)
            cell.font = label_font
        
        ws['A18'] = 'Nominal Value'
        ws['B18'] = '-'
        ws['C18'] = "='Sensor Specification'!C27"
        ws['D18'] = "='Sensor Specification'!D27"
        ws['E18'] = "='Sensor Specification'!E27"
        ws['F18'] = "='Sensor Specification'!F27"
        ws['G18'] = "='Sensor Specification'!G27"
        ws['H18'] = 'yes' if self.specs.get('Apodized', 'Yes').lower() == 'yes' else 'no'
        
        ws['A19'] = 'Plus Tolerance'
        ws['B19'] = "='Sensor Specification'!B28"
        ws['C19'] = "='Sensor Specification'!C28"
        ws['D19'] = "='Sensor Specification'!D28"
        ws['E19'] = '-'
        ws['F19'] = "='Sensor Specification'!F28"
        
        ws['A20'] = 'Minus Tolerance'
        ws['B20'] = "='Sensor Specification'!B29"
        ws['C20'] = "='Sensor Specification'!C29"
        ws['D20'] = "='Sensor Specification'!D29"
        ws['E20'] = '-'
        ws['F20'] = "='Sensor Specification'!F29"
        
        # Fiber Design section
        ws['A22'] = 'Fiber Design'
        ws['A22'].font = header_font
        ws['A22'].fill = header_fill
        
        # Additional Specifications
        ws['A25'] = 'Additional Specifications'
        ws['A25'].font = header_font
        ws['A25'].fill = header_fill
        
        add_headers = ['Spectrum Datasheet', 'Label', 'Total Fiber Length (mm)', 
                      'Simplified Packaging', 'Tolerance FBG Spacing (%)', 'FBG Marking', '', 'Fiber Input Marking']
        for col, header in enumerate(add_headers, 1):
            cell = ws.cell(row=26, column=col, value=header)
            cell.font = label_font
        
        ws['A27'] = self.specs.get('Spectrum Datasheet', 'linear')
        ws['B27'] = "='Sensor Specification'!B34"
        ws['C27'] = "='Sensor Specification'!C34"
        ws['D27'] = 'no'
        ws['E27'] = 1
        ws['F27'] = 'black, FBG-length (1mm tolerance)'
        ws['H27'] = 'black'
        
        # Wavelength Table section
        ws['A29'] = 'Wavelength Table'
        ws['A29'].font = header_font
        ws['A29'].fill = header_fill
        
        wl_headers = ['FBG Name', 'Wavelength (nm)', 'FBG Position (mm)', 'FBG Lead Out (mm)', 'FBG Spacing (mm)']
        for col, header in enumerate(wl_headers, 1):
            cell = ws.cell(row=30, column=col, value=header)
            cell.font = label_font
        
        # Link to Sensor Specification wavelength table
        for i in range(20):
            row = 31 + i
            ws.cell(row=row, column=1, value=f"=IF(ISBLANK('Sensor Specification'!A{38+i}),\"\",'Sensor Specification'!A{38+i})")
            ws.cell(row=row, column=2, value=f"=IF(ISNUMBER('Sensor Specification'!B{38+i}),'Sensor Specification'!B{38+i},\"\")")
            ws.cell(row=row, column=3, value=f"=IF(ISNUMBER('Sensor Specification'!C{38+i}),'Sensor Specification'!C{38+i},\"\")")
            ws.cell(row=row, column=4, value=f"=IF(ISNUMBER('Sensor Specification'!D{38+i}),'Sensor Specification'!D{38+i},\"\")")
            ws.cell(row=row, column=5, value='-' if i == 0 else f'=IF(ISNUMBER(C{row}),C{row}-C{row-1},"")')
        
        # Footnotes
        ws['G53'] = '1) continuous fiber with multiple FBGs/arrays, to be coiled before delivery'
        ws['G55'] = '2) min. Tolerance Spacing: 1 mm'
        
        # Signature area
        ws['A58'] = 'Date:'
        ws['C58'] = "Customer's Signature:"
        
        # Column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 18
    
    def _build_definitions_sheet(self, ws, header_font, label_font, value_font, border):
        """Build the Definitions sheet with lookup data"""
        # Contact names
        ws['A1'] = 'Contact'
        ws['A1'].font = header_font
        contacts = ['Friedrich', 'Gillooly', 'Gong', 'Kampling', 'Reif', 'Vasileiou', 'Vouilleme', 'Yang', 'Hoffmann']
        for i, contact in enumerate(contacts, 2):
            ws[f'A{i}'] = contact
        
        # Fiber types lookup table
        headers = ['Fiber Type', 'Core Type', 'Mode Field', 'Fiber Cladding', 'Coating', 'Manufacturer', 'ITU-T']
        for col, header in enumerate(headers, 3):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = header_font
        
        fiber_types = [
            ('SMF-28 Ultra', 'Single Mode', 10.4, 125, 'Acrylate', 'Corning', 'G.657.A1'),
            ('SMF-28 Ultra 200', 'Single Mode', 10.4, 125, 'Acrylate', 'Corning', 'G.657.A1'),
            ('SM1330-E9/125PI', 'Single Mode', '10,4', '125', 'Polyimide', 'J-Fiber', '-'),
            ('SM1500SC(7/80)P', 'Pure Core', '6,7 - 7,6', '80', 'Polyimide', 'Fibercore', '-'),
            ('SM1250BI(9.8/125)P', 'Bend Insensitive', '9 - 10,6', '125', 'Polyimide', 'Fibercore', '-'),
            ('SM1250SC(9/125)', 'Pure Core', '8,3 - 9,6', '125', 'Acrylate', 'Fibercore', '-'),
            ('SM1250SC(9/125)P', 'Pure Core', '8,3 - 9,6', '125', 'Polyimide', 'Fibercore', '-'),
            ('SM1250SC(10/125)P', 'Pure Core', '9,2 - 10,8', 125, 'Polyimide', 'Fibercore', '-'),
            ('IXF-RAD-SM-1550-014-PI', 'Radiation Hard', '8 - 10', '125', 'Polyimide', 'iXblue', '-'),
            ('SMF-60-MTDA-125-1', 'Pure Core', '9,3 - 10,7', '125', 'Acrylate', 'Verrillon', '-'),
            ('Ormocer', 'Single Mode', '-', '125', 'Ormocer', 'J-Fiber', '-'),
            ('Single Mode Acrylate', 'Single Mode', '-', '-', 'Acrylate', '-', '-'),
            ('Single Mode PI', 'Single Mode', '-', '-', 'Polyimide', '-', '-'),
            ('Pure Core PI', 'Pure Core', '-', '-', 'Polyimide', '-', '-'),
            ("Other (see Remarks)", '-', '-', '-', '-', '-', '-'),
        ]
        
        for row_idx, fiber in enumerate(fiber_types, 2):
            for col_idx, value in enumerate(fiber):
                ws.cell(row=row_idx, column=col_idx + 3, value=value)
        
        # Provided By options
        ws['K1'] = 'Provided'
        ws['K1'].font = header_font
        ws['K2'] = 'eFG'
        ws['K3'] = 'customer'
        
        # FemtoPlus options
        ws['M1'] = 'FemtoPlus'
        ws['M1'].font = header_font
        ws['M2'] = 'yes'
        ws['M3'] = 'no'
        
        # Column widths
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
    
    def _build_drawings_sheet(self, ws, header_font, label_font):
        """Build the Drawings sheet with configuration options"""
        ws['A3'] = 'Capillary without Tube'
        ws['A4'] = 'Capillary with Tube'
        ws['A5'] = 'Fiber Only'
        ws['A6'] = 'Fiber with Tube'
        
        # Drawing reference formula
        ws['B1'] = "=IF('Sensor Specification'!A18=Drawings!A3,3,IF('Sensor Specification'!A18=Drawings!A4,4,IF('Sensor Specification'!A18=Drawings!A5,5,IF('Sensor Specification'!A18=Drawings!A6,6,0))))"
        
        ws.column_dimensions['A'].width = 25
    
    def generate_pdf_quote(self, output_path: str) -> str:
        """Generate PDF quote document matching the template format"""
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            leftMargin=20*mm,
            rightMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey
        )
        
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=10
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            spaceBefore=4,
            spaceAfter=4
        )
        
        small_style = ParagraphStyle(
            'Small',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.grey
        )
        
        story = []
        
        # Company header
        company_header = f"""
        <b>{self.company_info['name']}</b><br/>
        {self.company_info['address']}<br/>
        {self.company_info['city']}<br/>
        {self.company_info['phone']}<br/>
        {self.company_info['email']}<br/>
        {self.company_info['website']}
        """
        story.append(Paragraph(company_header, normal_style))
        story.append(Spacer(1, 10*mm))
        
        # Address line
        address_line = f"{self.company_info['name']} - {self.company_info['address']} - {self.company_info['city']} - {self.company_info['country']}"
        story.append(Paragraph(address_line, small_style))
        story.append(Spacer(1, 5*mm))
        
        # Customer info
        customer = self.specs.get('Customer', 'NRL - US Naval Research Laboratory')
        story.append(Paragraph(customer, normal_style))
        story.append(Paragraph('USA', normal_style))
        story.append(Spacer(1, 5*mm))
        
        # Right side info block
        date_str = datetime.now().strftime('%d.%m.%Y')
        info_data = [
            ['Page:', '1'],
            ['Customer No.:', '20797'],
            ['Order No.:', f'{self.pricing["quantity"]}x Conn. {int(self.specs.get("Number of FBGs", "2"))}FBG arrays'],
            ['Agent:', 'Andrew Gillooly'],
            ['Delivery date:', date_str],
            ['Date:', date_str],
        ]
        
        info_table = Table(info_data, colWidths=[80, 100])
        info_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 10*mm))
        
        # Offer title
        story.append(Paragraph(f"<b>Offer No. {self.offer_number}</b>", title_style))
        story.append(Spacer(1, 5*mm))
        
        # Greeting
        customer_short = customer.split(' - ')[0] if ' - ' in customer else customer
        story.append(Paragraph(f"Dear {customer_short},", normal_style))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(
            "We are happy to provide you the following non-binding offer subject to our general terms and conditions of delivery and payment:",
            normal_style
        ))
        story.append(Spacer(1, 5*mm))
        
        # Quote table
        qty = self.pricing['quantity']
        unit_price = self.pricing['unit_price']
        total_product = qty * unit_price
        customs = self.pricing['customs']
        net_amount = total_product + customs
        
        num_fbgs = int(self.specs.get('Number of FBGs', '2'))
        description = f"{num_fbgs} FBG arrays with FC/APC connectors on both ends.\nSee spec sheet {self.offer_number}"
        
        quote_data = [
            ['Item', 'Qty.', 'Unit', 'Art.-No.', 'Description', 'Unit price\nEUR', 'Value\nEUR'],
            ['1', str(qty), 'pcs.', '', description, f'{unit_price:.4f}', f'{total_product:,.2f}'],
        ]
        
        quote_table = Table(quote_data, colWidths=[30, 30, 30, 40, 150, 55, 55])
        quote_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('ALIGN', (5, 1), (6, -1), 'RIGHT'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(quote_table)
        
        # Totals
        totals_data = [
            ['plus Customs Declaration (Germany)', '', '', '', '', '', f'{customs:.2f}'],
            ['Net amount', '', '', '', '', '', f'{net_amount:,.2f}'],
            ['tax free (NON-EU)', '', '', '', '', f'{net_amount:,.2f}', ''],
            ['Total value', '', '', '', '', '', f'{net_amount:,.2f}'],
        ]
        
        totals_table = Table(totals_data, colWidths=[150, 30, 30, 40, 40, 55, 55])
        totals_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (5, 0), (6, -1), 'RIGHT'),
            ('LINEABOVE', (0, 3), (-1, 3), 1, colors.black),
            ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ]))
        story.append(totals_table)
        story.append(Spacer(1, 10*mm))
        
        # Terms
        terms = [
            "Terms of payment: Payment in advance",
            "Incoterms 2020: EXW",
            "",
            "Please return signed specification sheet with order.",
            "This offer is valid until: 30 days",
            "All prices are net prices excluding shipping costs, VAT, taxes, customs and other fees for exportation.",
            "Notices of defect are only accepted within 14 days after delivery.",
            "Goods with value > 5000€ will be shipped insured. Shipping costs will be borne by the customer.",
            "The commodities listed in this document are classified as dual-use items under Council Regulation (EC) No 2021/821 [Annex I; list number 6A002D] by The German Federal Office for Economic Affairs and Export Control (BAFA).",
            "Prospective Shipping timeframe: TBD"
        ]
        
        for term in terms:
            story.append(Paragraph(term, normal_style))
        
        story.append(Spacer(1, 20*mm))
        
        # Footer
        footer_data = [
            [self.company_info['name'], f"Managing Director: {self.company_info['managing_director']}", 
             f"Finance Office Goslar", 'Bank details:'],
            [self.company_info['court'], '', f"Tax No.{self.company_info['tax_no']}", self.company_info['bank']],
            [self.company_info['hrb'], f"EORI: {self.company_info['eori']}", f"VAT {self.company_info['vat']}", 
             f"IBAN: {self.company_info['iban']}"],
            ['', '', '', f"SWIFT-BIC: {self.company_info['swift']}"],
        ]
        
        footer_table = Table(footer_data, colWidths=[100, 100, 100, 120])
        footer_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey),
            ('LINEABOVE', (0, 0), (-1, 0), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, 0), 5),
        ]))
        story.append(footer_table)
        
        # Build PDF
        doc.build(story)
        return output_path


def generate_quote_documents(specifications: dict, output_dir: str, offer_number: str = "41260018"):
    """Main function to generate both quote documents"""
    os.makedirs(output_dir, exist_ok=True)
    
    generator = QuoteDocumentGenerator(specifications, offer_number)
    
    excel_path = os.path.join(output_dir, f"{offer_number}_NRL.xlsx")
    pdf_path = os.path.join(output_dir, f"{offer_number}_NRL.pdf")
    
    excel_file = generator.generate_excel_spec_sheet(excel_path)
    pdf_file = generator.generate_pdf_quote(pdf_path)
    
    return {
        'excel': excel_file,
        'pdf': pdf_file
    }


if __name__ == "__main__":
    # Test with sample specifications
    test_specs = {
        'Customer': 'NRL - US Naval Research Laboratory',
        'Customer Number': '20797',
        'Quantity': '10 pcs',
        'Configuration': '2 FBG arrays with FC/APC connectors on both ends',
        'Fiber Type': 'SM1330-E9/125PI',
        'Connector Type': 'FC/APC both ends',
        'Total Fiber Length': '10050 mm',
        'Number of FBGs': '2',
        'Fiber Coating': 'Polyimide',
        'FBG 1 Wavelength': '1550.39 nm',
        'FBG 2 Wavelength': '1555.39 nm',
        'Wavelength Tolerance': '±0.1 nm',
        'Reflectivity': '10%',
        'Reflectivity Tolerance': '±4%',
        'FWHM': '0.09 nm',
        'FWHM Tolerance': '±0.02 nm',
        'SLSR Minimum': '8 dB',
        'FBG Length': '12 mm',
        'FBG Length Tolerance': '±2 mm',
        'FBG Spacing': '50 mm',
        'First FBG Position': '5000 mm',
        'FemtoPlus': 'Yes',
        'Apodized': 'Yes',
        'Spectrum Datasheet': 'linear',
        'Label': 'on spool'
    }
    
    result = generate_quote_documents(test_specs, '/home/claude/rfq-outlook-addin/output')
    print(f"Generated documents:")
    print(f"  Excel: {result['excel']}")
    print(f"  PDF: {result['pdf']}")
